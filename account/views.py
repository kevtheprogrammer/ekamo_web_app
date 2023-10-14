from typing import Any
from django.shortcuts import render,redirect
from django.db.models import Q
from django.views.generic import ListView,DetailView, View
from transactions.forms import AgentExpensesForm, FilterForm
from transactions.models import AgentExpenses, FispTransaction
from datetime import datetime
from django.http import HttpResponse
import requests
from django.db.utils import IntegrityError
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP
from django.utils import timezone

from .models import *

import openpyxl


# List view for AgentProfile
class AgentProfileListView(ListView):
    model = AgentProfile
    template_name = 'control/agentprofile_list.html'
    context_object_name = 'agentprofiles'
    paginate_by = 15
    

    def get_context_data(self, **kwargs: Any):
        context =  super().get_context_data(**kwargs)
        context['total_agents'] = AgentProfile.objects.all().count()
        context['unique_provinces'] = AgentProfile.objects.values_list('province', flat=True).distinct()
        return context


# List view for AgentProfile
class AgentProfileFilterListView(ListView):
    model = AgentProfile
    template_name = 'control/agentprofile_list.html'
    context_object_name = 'agentprofiles'
    paginate_by = 30


    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('q')
        province = self.request.GET.get('province')

        if province == "all":
            return queryset

        if province:
            queryset = queryset.filter(province=province)

        if name:
            queryset = queryset.filter(
                Q(phonenumber__icontains=name) |
                Q(first_name__icontains=name) |
                Q(last_name__icontains=name) |
                Q(code__icontains=name) |
                Q(district__icontains=name) 
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        province = self.request.GET.get('province')
        name = self.request.GET.get('q')
        if province != "all":
            context["province"] = province
        if name:
            context["query"] = name
        context['total_agents'] = AgentProfile.objects.all().count()
        context['unique_provinces'] = AgentProfile.objects.values_list('province', flat=True).distinct()
        return context


class AgentProfileDetailView(DetailView):
    model = AgentProfile
    template_name = 'control/agentprofile_detail.html'
    context_object_name = 'agentprofiles'
    form_class = AgentExpensesForm
     
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs) 
        # Use the start_date and end_date in your transaction filtering logic
        txn = FispTransaction.objects.filter(agent=self.get_object())
        
        total_amount =  txn.aggregate(total_amount=models.Sum('transAmount'))['total_amount']
        div = Decimal(400)
        agent_com = Decimal(0)
        bill_com = Decimal(0)
        nat_com = Decimal(12) 
        com = Decimal(self.get_object().get_earning())

        if com == 0:
            # Handle the case where com is 0
            agent_com = Decimal(0)
            bill_com = (total_amount / div) * nat_com
        elif com > nat_com:
            agent_com = com
            bill_com = (total_amount / div) * nat_com
        else:
            agent_com = (total_amount / div) * com
            bill_com_unit = nat_com - com
            bill_com = (total_amount / div) * bill_com_unit


  
        exp =  AgentExpenses.objects.filter(agent_given=self.get_object())

        rounded_value_agent_com  = agent_com.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        rounded_value_bill_com  = bill_com.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        context["total_amount"] = total_amount
        context["total_transaction_count"] = txn.count()
        context["agent_commission"] =   rounded_value_agent_com
        context["bill_comm"] = rounded_value_bill_com
        context["transactions"] = txn 
        context["form"] = self.form_class()
        context["expenses"] = exp
        context["total_amount_expenses"] =  exp.aggregate(total_amount=models.Sum('amount'))['total_amount']
        context["total_amount_expenses_count"] =  exp.count()
       
        return context
    
    def post(self,request, *args, **kwargs):
        form = AgentExpensesForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.agent_given = self.get_object()
            expense.save()
            return redirect(self.get_object().get_absolute_url())


class AgentProfileFilterDetailView(DetailView):
    model = AgentProfile
    template_name = 'control/agentprofile_detail.html'
    context_object_name = 'agentprofiles'
    form_class = AgentExpensesForm
     
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('timestamp')
        end_date = self.request.GET.get('date_to')   
        # Use the start_date and end_date in your transaction filtering logic
        txn = FispTransaction.objects.filter(agent=self.get_object())
        
        if start_date == end_date:
            txn = txn.filter(created_at__date=start_date)
        else:
            if start_date:
                txn = txn.filter(created_at__gte=start_date)
            if end_date:
                txn = txn.filter(created_at__lte=end_date)

        total_amount =  txn.aggregate(total_amount=models.Sum('transAmount'))['total_amount']
        print('tital txn -->>>', txn,total_amount )
        div = Decimal(400)
        agent_com = Decimal(0)
        bill_com = Decimal(0)
        nat_com = Decimal(12) 
        com = Decimal(self.get_object().get_earning())

        try:
            if com == 0:
                # Handle the case where com is 0
                agent_com = Decimal(0)
                bill_com = (total_amount / div) * nat_com
            elif com > nat_com:
                agent_com = com
                bill_com = (total_amount / div) * nat_com
            else:
                agent_com = (total_amount / div) * com
                bill_com_unit = nat_com - com
                bill_com = (total_amount / div) * bill_com_unit
        except:
            pass

  
        exp =  AgentExpenses.objects.filter(agent_given=self.get_object())

        rounded_value_agent_com  = agent_com.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
        rounded_value_bill_com  = bill_com.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        context["start_date"] = start_date
        context["end_date"] = end_date
        context["total_amount"] = total_amount
        context["total_transaction_count"] = txn.count()
        context["agent_commission"] =   rounded_value_agent_com
        context["bill_comm"] = rounded_value_bill_com
        context["transactions"] = txn 
        context["form"] = self.form_class()
        context["expenses"] = exp
        context["total_amount_expenses"] =  exp.aggregate(total_amount=models.Sum('amount'))['total_amount']
        context["total_amount_expenses_count"] =  exp.count()
       
        return context
    
    def post(self,request, *args, **kwargs):
        form = AgentExpensesForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.agent_given = self.get_object()
            expense.save()
            return redirect(self.get_object().get_absolute_url())


def export_all_transactions_excel(request):
    # Create a new workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="all_transactions.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

    # Write header row
    header = ['ATD Number', 'Amount', 'Agent', 'Agent Number', 'RefId', 'Txn Time']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = FispTransaction.objects.all().values_list('transAdt', 'transAmount', 'agent__first_name', 'agent__last_name', 'agent__phonenumber', 'transRef', 'created_at')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, timezone.datetime):
                cell_value = cell_value.replace(tzinfo=None)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_all_agents_excel(request):
    # Create a new workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="all_agents.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Agents'

    # Write header row
    header = ['First Name', 'Last Name', 'Code','Agent Type' ,'Agent Number', 'Province', 'District']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = AgentProfile.objects.all().values_list('first_name', 'last_name', 'code', 'agent_type__title', 'phonenumber', 'province', 'district')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, timezone.datetime):
                cell_value = cell_value.replace(tzinfo=None)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_transactions_excel(request, pk):
    # Create a new workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

    # Write header row
    header = ['ATD Number', 'Amount', 'Agent', 'Agent Number', 'RefId', 'Txn Time']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = FispTransaction.objects.filter(agent__id=pk).values_list('transAdt', 'transAmount', 'agent__first_name', 'agent__last_name', 'agent__phonenumber', 'transRef', 'created_at')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, timezone.datetime):
                cell_value = cell_value.replace(tzinfo=None)
            cell.value = cell_value

    workbook.save(response)

    return response

def export_agent_expenses_excel(request, pk):
    # Create a new workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="AgentExpenses.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Agent Expenses' 

    # Write header row
    header = [ 'Amount', 'Agent First Name', 'Agent Last Name', 'Agent Number', 'Reason', 'Time']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = AgentExpenses.objects.filter(agent_given__id=pk).values_list('amount','agent_given__first_name', 'agent_given__last_name', 'agent_given__phonenumber','reason',  'timestamp')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if isinstance(cell_value, timezone.datetime):
                cell_value = cell_value.replace(tzinfo=None)
            cell.value = cell_value
    workbook.save(response)

    return response


def detail_search_txn(request, pk):
    transactions = FispTransaction.objects.filter(agent__pk=pk)
    
    # Initialize variables for search and total amount
    search_query = ""
    total_amount = None

    if request.method == "POST":
        # Retrieve the search query from the POST data
        search_query = request.POST.get('q')

        # Filter transactions based on search query
        transactions = transactions.filter(
            Q(transAdt__icontains=search_query) |
            Q(transAmount__icontains=search_query) |
            Q(leadernumber__icontains=search_query) |
            Q(depositer__icontains=search_query) |
            Q(agent__first_name__icontains=search_query) |
            Q(agent__last_name__icontains=search_query)
        )

        # Calculate the total amount for the filtered transactions
        total_amount = transactions.aggregate(total_amount=models.Sum('transAmount'))['total_amount']

        deposited_amount = transactions.aggregate(total_amount=Sum('transAmount'))['total_amount']
        print('total_amount',total_amount)
        print('deposited_amount',deposited_amount)
        # deposited_amount = transactions.aggregate(total_amount=models.Sum('transAmount'))['total_amount']
        # deposited_await = total_amount - deposited_amount 
        context = {
            'form': FilterForm(user=request.user),
            'total_amount': total_amount,
            'transactions':transactions,
            # 'my_agents':agents_to_manage,
            'deposited_amount':deposited_amount ,
            'deposited_await':'deposited_await',
            'total_transaction_count': transactions.count(),
        }

        return render(request, 'control/transactions_search.html', context)
    else:
        form = FispTransaction()

    return render(request, 'control/transactions_search.html', {
        'transactions': transactions,
    })


def fetchAPI(request):
    # API endpoint
    base_url = "https://fisp.ekamo.co.zm/api/fisp/transactions"
    
    # Initialize variables
    page = 1
    all_data_fetched = False
    while not all_data_fetched:
        # Construct the URL for the current page
        url = f"{base_url}?page={page}"

        # Send a GET request to the API
        response = requests.get(url)

        if response.status_code == 200:
            api_data = response.json()
            transactions_data = api_data.get("data", [])

            if not transactions_data:
                # No more data on this page, exit the loop
                all_data_fetched = True
                continue

            for transaction_item in transactions_data:
                datetime_string  =  transaction_item.get("created_at", "")
                # Extract data from the API response
                trans_adt = transaction_item.get("transAdt", "")
                number_of_farmers = transaction_item.get("numberOfFarmers", 0)
                leadernumber = transaction_item.get("leadernumber", 0)  # Correct field name
                depositer = transaction_item.get("depositer", 0)  # Correct field name
                transRef = transaction_item.get("transRef", 0)
                totalCommis = transaction_item.get("totalCommis", 0)
                transCommisAgent = transaction_item.get("transCommisAgent", 0)
                transNewBalance = transaction_item.get("transNewBalance", 0)
                transOldComBalance = transaction_item.get("transOldComBalance", 0)
                transNewComBalance = transaction_item.get("transNewComBalance", 0)
                transStatus = transaction_item.get("transStatus", "")  # Correct field type
                status = transaction_item.get("status", 0)
                transAmount = transaction_item.get("transAmount", 0)
                created_at = datetime.strptime(datetime_string, "%d-%m-%Y %H:%M:%S")

                status_boolean = False
                if status == 1:  # Example: Map 1 to True, everything else to False
                    status_boolean = True
                else:
                    status_boolean = False

                transOldBalance = transaction_item.get("transOldBalance", 0)
                phone_number = transaction_item.get("phoneNumber", "")  # Add phone_number

                # Check if an agent with the given phone number exists, or create a new one
                agent, created = AgentProfile.objects.get_or_create(
                    phonenumber=phone_number,
                    defaults={
                        'floatLimit': 1000.00,  # Default float limit for new agents
                        # You can set other default values here
                    }
                )

                # Check if the transaction has a successful status before saving it
                if transStatus == "success":
                    # Create and save a new FispTransaction instance
                    transaction = FispTransaction(
                        transAdt=trans_adt,
                        numberOfFarmers=number_of_farmers,
                        transAmount=transAmount,
                        transOldBalance=transOldBalance,
                        leadernumber=leadernumber,
                        depositer=depositer,
                        transRef=transRef,
                        totalCommis=totalCommis,
                        transCommisAgent=transCommisAgent,
                        transNewBalance=transNewBalance,
                        transOldComBalance=transOldComBalance,
                        transNewComBalance=transNewComBalance,
                        created_at=created_at,
                        isSuccess=status_boolean,       
                        agent=agent  # Use the agent as the primary key for the transaction
                        # Set other fields accordingly
                    )
                    try:
                        transaction.save()
                    except IntegrityError:
                        # Duplicate transRef, skip this entry and continue with the next one
                        pass
            # Move to the next page
            page += 1
    return redirect('account:txn')



  